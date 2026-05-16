import streamlit as st
import pandas as pd
import numpy as np
import io
import pulp
import traceback
import random
import openpyxl
from datetime import datetime

# ==========================================
# 1. 網頁頁面配置
# ==========================================
st.set_page_config(page_title="段考監考終極自動化", page_icon="🏫", layout="wide")
st.title("🏫 試務組-段考監考全自動化系統 (動態自適應版)")
st.info("💡 終極升級：系統已解除固定 10 節的限制！現在 AI 會根據您上傳的檔案，自動偵測半天/全天的節次總數與換日線，完美適應 8 節、10 節等各種排班情境。")

# --- 初始化狀態 ---
if 'results' not in st.session_state:
    st.session_state['results'] = None
if 'uploader_key' not in st.session_state:
    st.session_state['uploader_key'] = 0

# ==========================================
# 2. 輔助功能定義
# ==========================================
def to_excel_bytes(df, header_df=None):
    output = io.BytesIO()
    if header_df is not None:
        df.columns = header_df.columns
        final_out = pd.concat([header_df, df], ignore_index=True)
    else:
        final_out = df
    
    final_out = final_out.fillna("")
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        final_out.to_excel(writer, index=False, header=False)
    return output.getvalue()

def normalize_cls(c):
    s = str(c).strip().replace('ㄧ', '一').replace(' ', '').replace('　', '')
    s = s.translate(str.maketrans('１２３４５６７８９０', '1234567890'))
    return s

def normalize_subject(s):
    s = str(s).strip().replace(' ', '').replace('　', '')
    s = s.replace('國文', '國語文').replace('英文', '英語文')
    return s

# ==========================================
# 3. 介面佈局
# ==========================================
st.divider()
col1, col2 = st.columns([1, 1], gap="large")

with col1:
    st.subheader("📂 1. 上傳排考與標籤資料")
    file_quota = st.file_uploader("1️⃣ 監考堂數.xlsx", type=['xlsx'], key=f"f1_{st.session_state['uploader_key']}")
    file_list = st.file_uploader("2️⃣ 監考名單.xlsx", type=['xlsx'], key=f"f2_{st.session_state['uploader_key']}")
    file_type = st.file_uploader("3️⃣ 監考類型總數.xlsx", type=['xlsx'], key=f"f3_{st.session_state['uploader_key']}")
    file_pub = st.file_uploader("4️⃣ 監考總表公布版.xlsx (範本)", type=['xlsx'], key=f"f4_{st.session_state['uploader_key']}")
    file_assign = st.file_uploader("5️⃣ 監考一覽表.xlsx (手排前導日完成版)", type=['xlsx'], key=f"f5_{st.session_state['uploader_key']}")
    st.write("---")
    file_course = st.file_uploader("6️⃣ 配課表.xlsx (多工作表)", type=['xlsx'], key=f"f6_{st.session_state['uploader_key']}")
    file_label = st.file_uploader("7️⃣ 標籤列印.xlsx (試卷袋範本)", type=['xlsx'], key=f"f7_{st.session_state['uploader_key']}")

with col2:
    st.subheader("⚙️ 2. 考試設定與特許名單")
    selected_sheet = None
    if file_quota:
        xls = pd.ExcelFile(file_quota)
        selected_sheet = st.selectbox("👇 選擇考試項目：", xls.sheet_names)
    
    flex_names = []
    if file_list:
        temp_df = pd.read_excel(file_list, header=None).fillna("")
        teacher_list = temp_df.iloc[2:, 1].astype(str).str.strip().tolist()
        teacher_list = [t for t in teacher_list if t != "" and t != "nan"]
        flex_names = st.multiselect("🛡️ 優先時數不大於名單：", options=teacher_list)

    st.write("")
    c_d0, c_d1, c_d2 = st.columns(3)
    with c_d0: d0_date = st.date_input("📅 手排日(如13日)", datetime.now())
    with c_d1: d1_date = st.date_input("📅 AI Day1(如14日)", datetime.now())
    with c_d2: d2_date = st.date_input("📅 AI Day2(如15日)", datetime.now())
    
    force_run = st.checkbox("⚠️ 忽略健檢警告，強制執行")
    if st.button("🗑️ 清除所有設定", use_container_width=True):
        st.session_state['results'] = None
        st.session_state['uploader_key'] += 1
        st.rerun()

# ==========================================
# 4. 核心演算法執行
# ==========================================
st.divider()

if st.button("🚀 啟動終極全自動排班系統", type="primary", use_container_width=True):
    if not all([file_quota, file_list, file_type, file_assign]):
        st.error("🚨 請至少確認【1, 2, 3, 5】號基礎檔案皆已上傳！")
    else:
        try:
            # --- 讀取基礎資料 ---
            df_quota = pd.read_excel(file_quota, sheet_name=selected_sheet).fillna("")
            quota_dict = dict(zip(df_quota.iloc[:, 0].astype(str).str.strip(), pd.to_numeric(df_quota.iloc[:, 1], errors='coerce').fillna(0)))
            
            df_list_raw = pd.read_excel(file_list, header=None).fillna("")
            
            # 【核心動態偵測一】：掃描名單總共有幾節課，不再預設 11 節
            period_cols = []
            for c in range(1, df_list_raw.shape[1]):
                val = str(df_list_raw.iloc[1, c]).strip()
                try:
                    int(float(val))
                    period_cols.append(c)
                except: pass
            
            total_periods = len(period_cols)
            if total_periods < 2:
                st.error("🚨 無法從檔案中辨識出節次數量，請檢查監考名單格式。")
                st.stop()
                
            ai_periods = total_periods - 1  # 扣掉第一欄 (手動排的 13日)
            all_period_nums = [int(float(str(df_list_raw.iloc[1, c]).strip())) for c in period_cols]
            ai_period_nums = all_period_nums[1:]
            
            # 【核心動態偵測二】：尋找「換日線」 (當節次號碼變小，代表換天了)
            day_starts = [0]
            for j in range(1, ai_periods):
                if ai_period_nums[j] <= ai_period_nums[j-1]:
                    day_starts.append(j)

            # 動態建立 req_matrix (根據實際總節次讀取人數)
            df_type = pd.read_excel(file_type, header=None).fillna("")
            req_matrix = {'△': [0]*ai_periods, '※': [0]*ai_periods}
            for i in range(2, len(df_type)):
                row_name = str(df_type.iloc[i, 0]).strip()
                if row_name in ['△', '※']:
                    req_list = pd.to_numeric(df_type.iloc[i, 1:], errors='coerce').fillna(0).astype(int).tolist()
                    if len(req_list) < total_periods:
                        req_list += [0] * (total_periods - len(req_list))
                    req_matrix[row_name] = req_list[1 : 1 + ai_periods]

            d0_str, d1_str, d2_str = d0_date.strftime('%m月%d日'), d1_date.strftime('%m月%d日'), d2_date.strftime('%m月%d日')
            
            header_df = df_list_raw.iloc[0:2].copy().astype(str).replace('nan', '')
            header_df.iloc[0, period_cols[0]] = d0_str
            for j in range(ai_periods):
                d_str = d1_str
                if len(day_starts) > 1 and j >= day_starts[1]: d_str = d2_str
                header_df.iloc[0, period_cols[j+1]] = d_str
            
            df_list = df_list_raw.iloc[2:].copy()
            teachers = df_list.iloc[:, 1].astype(str).str.strip().tolist()

            # --- PuLP 運算 (完全動態自適應) ---
            with st.spinner(f"🧠 AI 自動偵測到 {ai_periods} 節需排班，正在生成完美監考總表..."):
                prob = pulp.LpProblem("Scheduling", pulp.LpMinimize)
                vX = {}; vY = {}
                for i in range(len(teachers)):
                    vX[i] = {}; vY[i] = {}
                    for j in range(ai_periods):
                        vX[i][j] = pulp.LpVariable(f"X_{i}_{j}", cat='Binary')
                        vY[i][j] = pulp.LpVariable(f"Y_{i}_{j}", cat='Binary')
                
                penalty = 0
                for i, t in enumerate(teachers):
                    tgt = int(quota_dict.get(t, 0))
                    act = pulp.lpSum([vX[i][k] + vY[i][k]*2 for k in range(ai_periods)])
                    prob += act <= tgt
                    dfct = pulp.LpVariable(f"dfct_{i}", 0)
                    prob += act + dfct == tgt
                    penalty += dfct * (1 if t in flex_names else 1000)
                    
                    for j in range(ai_periods):
                        prob += vX[i][j] + vY[i][j] <= 1
                        cell_val = str(df_list.iloc[i, period_cols[j+1]]).strip()
                        if cell_val != "" and cell_val != "nan":
                            prob += vX[i][j] == 0; prob += vY[i][j] == 0
                            
                    # 【核心動態偵測三】：動態連堂防呆，自動避開跨日連堂
                    for j in range(1, ai_periods):
                        if j not in day_starts:
                            prob += vX[i][j] >= vY[i][j-1]
                        else:
                            prob += vY[i][j-1] == 0  # 換日的前一節不允許排連堂
                    prob += vY[i][ai_periods-1] == 0
                    
                for j in range(ai_periods):
                    prob += pulp.lpSum([vX[i][j] for i in range(len(teachers))]) == req_matrix['△'][j]
                    prob += pulp.lpSum([vY[i][j] for i in range(len(teachers))]) == req_matrix['※'][j]
                prob += penalty
                prob.solve()

                schedule_dict = {}
                df_out_master = df_list.copy()
                for i, t in enumerate(teachers):
                    res = []
                    for j in range(ai_periods):
                        val = str(df_list.iloc[i, period_cols[j+1]]).strip()
                        if val == "" or val == "nan":
                            if vX[i][j].varValue == 1: val = "△"
                            elif vY[i][j].varValue == 1: val = "※"
                            else: val = "" 
                        res.append(val)
                        df_out_master.iloc[i, period_cols[j+1]] = val
                    schedule_dict[t] = res

            # --- 監考一覽表分配邏輯 ---
            with st.spinner("🎯 執行班級自動分配..."):
                df_assign_calc = pd.read_excel(file_assign, header=None).fillna("")
                raw_list = df_assign_calc.iloc[:, 0].astype(str).str.strip().tolist()
                class_names_raw = [x for x in raw_list if x and not any(bad in x for bad in ["班級", "日期", "節次", "星期", "一覽表", "總表", "華南", "期中考"])]
                
                assigned_matrix = np.empty((len(class_names_raw), ai_periods), dtype=object)
                
                # 依照動態偵測到的換日線進行分配
                for i_day, day_start in enumerate(day_starts):
                    day_end = day_starts[i_day+1] if i_day+1 < len(day_starts) else ai_periods
                    day_length = day_end - day_start
                    
                    j1 = day_start
                    proctors_j1 = [t for t in teachers if schedule_dict[t][j1] in ["△", "※"]]
                    random.shuffle(proctors_j1)
                    for idx, p in enumerate(proctors_j1): 
                        if idx < len(class_names_raw): assigned_matrix[idx, j1] = p
                    
                    if day_length > 1:
                        j2 = day_start + 1
                        proctors_j2 = [t for t in teachers if schedule_dict[t][j2] in ["△", "※"]]
                        bound = {}
                        for idx in range(len(class_names_raw)):
                            p_prev = assigned_matrix[idx, j1]
                            if p_prev is not None and p_prev in schedule_dict:
                                if schedule_dict[p_prev][j1] == "※" and schedule_dict[p_prev][j2] == "△":
                                    assigned_matrix[idx, j2] = p_prev
                                    bound[p_prev] = True
                        rem = [p for p in proctors_j2 if p not in bound]
                        random.shuffle(rem)
                        r_idx = 0
                        for idx in range(len(class_names_raw)):
                            if assigned_matrix[idx, j2] is None and r_idx < len(rem):
                                assigned_matrix[idx, j2] = rem[r_idx]; r_idx += 1

                        for offset in range(2, day_length):
                            curr_j = day_start + offset
                            proctors = [t for t in teachers if schedule_dict[t][curr_j] in ["△", "※"]]
                            random.shuffle(proctors)
                            for idx, p in enumerate(proctors): 
                                if idx < len(class_names_raw): assigned_matrix[idx, curr_j] = p

                class_proctor_schedule = {} 
                for r_idx, c_name in enumerate(class_names_raw):
                    class_proctor_schedule[normalize_cls(c_name)] = [assigned_matrix[r_idx, col] for col in range(ai_periods)]

                # --- 寫入一覽表 (動態適應 8~10 節) ---
                wb_assign = openpyxl.load_workbook(file_assign)
                ws_assign = wb_assign.active
                
                manual_day0_proctors = {}
                first_class_row = -1
                class_col_idx = 1
                for r in range(1, 20):
                    for c in range(1, 5):
                        v = ws_assign.cell(row=r, column=c).value
                        if v and str(v).strip() in class_names_raw:
                            first_class_row = r; class_col_idx = c; break
                    if first_class_row != -1: break
                
                if first_class_row != -1:
                    temp_cols = []
                    for r in range(max(1, first_class_row - 3), first_class_row + 1):
                        t_cols = []
                        for c in range(class_col_idx + 1, ws_assign.max_column + 1):
                            val = str(ws_assign.cell(row=r, column=c).value).strip()
                            try:
                                int(float(val))
                                t_cols.append(c)
                            except: pass
                        if len(t_cols) >= total_periods:
                            temp_cols = t_cols[:total_periods]; break
                    
                    if len(temp_cols) == total_periods:
                        date_row = first_class_row - 2
                        if date_row >= 1:
                            try: ws_assign.cell(row=date_row, column=temp_cols[0]).value = d0_str
                            except AttributeError: pass
                            
                            # 動態填寫日期
                            for j in range(ai_periods):
                                d_str = d1_str
                                if len(day_starts) > 1 and j >= day_starts[1]: d_str = d2_str
                                try: ws_assign.cell(row=date_row, column=temp_cols[j+1]).value = d_str
                                except AttributeError: pass
                        
                        for r in range(first_class_row, ws_assign.max_row + 1):
                            c_val = ws_assign.cell(row=r, column=class_col_idx).value
                            if c_val:
                                norm_c = normalize_cls(c_val)
                                val_manual = ws_assign.cell(row=r, column=temp_cols[0]).value
                                manual_day0_proctors[norm_c] = str(val_manual).strip() if val_manual else ""
                                
                                if norm_c in class_proctor_schedule:
                                    for j in range(ai_periods):
                                        ws_assign.cell(row=r, column=temp_cols[j+1]).value = class_proctor_schedule[norm_c][j]
                
                out_assign = io.BytesIO()
                wb_assign.save(out_assign)
                assign_bytes = out_assign.getvalue()

            # --- 公布版套印 (同理動態適應) ---
            pub_bytes = None
            if file_pub:
                with st.spinner("🖨️ 正在將資料無縫套印至公布版..."):
                    wb = openpyxl.load_workbook(file_pub)
                    ws = wb.active
                    h_row = -1; t_cols = []
                    for r in range(1, 16):
                        for c in range(1, 61):
                            val = ws.cell(row=r, column=c).value
                            if val and "教師" in str(val): h_row = r; t_cols.append(c)
                        if len(t_cols) > 0: break
                        
                    if h_row != -1:
                        for c in t_cols:
                            t_col_target = []
                            for scan_c in range(c + 1, c + 20):
                                val = str(ws.cell(row=h_row, column=scan_c).value).strip()
                                try:
                                    int(float(val))
                                    t_col_target.append(scan_c)
                                except: pass
                            
                            if len(t_col_target) >= total_periods:
                                try: ws.cell(row=h_row-1, column=t_col_target[0]).value = d0_str
                                except AttributeError: pass
                                
                                for j in range(ai_periods):
                                    d_str = d1_str
                                    if len(day_starts) > 1 and j >= day_starts[1]: d_str = d2_str
                                    try: ws.cell(row=h_row-1, column=t_col_target[j+1]).value = d_str
                                    except AttributeError: pass
                                
                                for r in range(h_row+1, ws.max_row + 1):
                                    t_val = ws.cell(row=r, column=c).value
                                    if t_val:
                                        name = str(t_val).strip()
                                        if name in schedule_dict:
                                            for j in range(ai_periods):
                                                ws.cell(row=r, column=t_col_target[j+1]).value = schedule_dict[name][j]
                    out_pub = io.BytesIO()
                    wb.save(out_pub)
                    pub_bytes = out_pub.getvalue()

            # --- 標籤列印自動生成邏輯 ---
            label_bytes = None
            if file_course and file_label:
                with st.spinner("🏷️ 正在合成試卷袋標籤..."):
                    course_dict = {}
                    xls_course = pd.ExcelFile(file_course)
                    for sheet in xls_course.sheet_names:
                        df_c = pd.read_excel(file_course, sheet_name=sheet).fillna("")
                        for r_idx, row in df_c.iterrows():
                            subj_raw = str(row.iloc[0]).strip()
                            if not subj_raw: continue
                            subj_norm = normalize_subject(subj_raw)
                            for c_idx in range(1, len(df_c.columns)):
                                cls_raw = str(df_c.columns[c_idx]).strip()
                                teacher = str(row.iloc[c_idx]).strip()
                                if teacher and cls_raw:
                                    course_dict[(normalize_cls(cls_raw), subj_norm)] = teacher
                    
                    wb_label = openpyxl.load_workbook(file_label)
                    ws_label = wb_label.active
                    col_map = {}
                    header_row = 1
                    for r in range(1, 6):
                        for c in range(1, ws_label.max_column + 1):
                            val = str(ws_label.cell(row=r, column=c).value).strip()
                            if "班級" in val and '班級' not in col_map: col_map['班級'] = c
                            elif "科目" in val and '科目' not in col_map: col_map['科目'] = c
                            elif "日期" in val and '日期' not in col_map: col_map['日期'] = c
                            elif "序號" in val and '序號' not in col_map: col_map['序號'] = c
                            elif "任課" in val and '任課教師' not in col_map: col_map['任課教師'] = c
                            elif "監考" in val and '監考老師' not in col_map: col_map['監考老師'] = c
                        if '班級' in col_map and '監考老師' in col_map:
                            header_row = r; break

                    d0_ymd, d0_short, d0_slash = d0_date.strftime('%Y-%m-%d'), d0_date.strftime('%m-%d'), d0_date.strftime('%Y/%m/%d')
                    d1_ymd, d1_short, d1_slash = d1_date.strftime('%Y-%m-%d'), d1_date.strftime('%m-%d'), d1_date.strftime('%Y/%m/%d')
                    d2_ymd, d2_short, d2_slash = d2_date.strftime('%Y-%m-%d'), d2_date.strftime('%m-%d'), d2_date.strftime('%Y/%m/%d')

                    # 動態建立節次對應表
                    day_p_val_to_ai_col = {}
                    curr_day_idx = 0
                    for j in range(ai_periods):
                        if j in day_starts and j != 0: curr_day_idx += 1
                        day_p_val_to_ai_col[(curr_day_idx, ai_period_nums[j])] = j

                    for r in range(header_row + 1, ws_label.max_row + 1):
                        if '班級' not in col_map: continue
                        cls_raw = ws_label.cell(row=r, column=col_map['班級']).value
                        if cls_raw is None or not str(cls_raw).strip(): continue
                        
                        subj_raw = ws_label.cell(row=r, column=col_map['科目']).value if '科目' in col_map else ""
                        date_val = ws_label.cell(row=r, column=col_map['日期']).value if '日期' in col_map else ""
                        date_str = str(date_val).split()[0].strip() if date_val is not None else ""
                        seq_val = ws_label.cell(row=r, column=col_map['序號']).value if '序號' in col_map else ""
                        
                        cls = normalize_cls(cls_raw)
                        subj = normalize_subject(subj_raw)
                        
                        if '任課教師' in col_map:
                            teacher = course_dict.get((cls, subj), "")
                            if not teacher:
                                for (c, s), t in course_dict.items():
                                    if c == cls and (subj in s or s in subj):
                                        teacher = t; break
                            if teacher: ws_label.cell(row=r, column=col_map['任課教師']).value = teacher
                        
                        try: p_val = int(float(str(seq_val).strip()))
                        except: p_val = -1
                        
                        if '監考老師' in col_map:
                            if any(d in date_str for d in [d0_ymd, d0_short, d0_slash]):
                                if cls in manual_day0_proctors:
                                    ws_label.cell(row=r, column=col_map['監考老師']).value = manual_day0_proctors[cls]
                            
                            elif cls in class_proctor_schedule and p_val != -1:
                                day_idx = -1
                                if any(d in date_str for d in [d1_ymd, d1_short, d1_slash]): day_idx = 0
                                elif any(d in date_str for d in [d2_ymd, d2_short, d2_slash]): day_idx = 1
                                
                                if day_idx != -1 and (day_idx, p_val) in day_p_val_to_ai_col:
                                    target_col = day_p_val_to_ai_col[(day_idx, p_val)]
                                    ws_label.cell(row=r, column=col_map['監考老師']).value = class_proctor_schedule[cls][target_col]

                    out_label = io.BytesIO()
                    wb_label.save(out_label)
                    label_bytes = out_label.getvalue()

            st.balloons()
            st.session_state['results'] = {
                'orig': to_excel_bytes(df_out_master, header_df),
                'assign': assign_bytes,
                'pub': pub_bytes,
                'label': label_bytes
            }

        except Exception as e:
            st.error(f"發生錯誤: {e}")
            st.code(traceback.format_exc())

# ==========================================
# 5. 下載區
# ==========================================
if st.session_state['results']:
    st.divider()
    res = st.session_state['results']
    c1, c2, c3, c4 = st.columns(4)
    with c1: st.download_button("📥 1. 監考總表", res['orig'], "監考總表.xlsx", "application/vnd.ms-excel", use_container_width=True)
    with c2: st.download_button("📥 2. 監考一覽表(保留手排)", res['assign'], "監考一覽表_分配完成.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, type="primary")
    with c3: 
        if res['pub']: st.download_button("📥 3. 公布版套印總表", res['pub'], "公布版總表.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with c4:
        if res.get('label'): st.download_button("📥 4. 標籤列印(完美接合)", res['label'], "標籤列印_完整版.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, type="primary")
